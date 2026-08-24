from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="B0B0B0")] * 4)


class ReportError(Exception):
    pass


class SourceNotFoundError(ReportError):
    pass


class ColumnMismatchError(ReportError):
    pass


@dataclass(frozen=True)
class ColumnSpec:
    name: str
    source_header: str
    width: int = 14
    number_format: str | None = None


@dataclass(frozen=True)
class Aggregation:
    column: str
    function: str
    label: str = "Total"

    def compute(self, values: list[float]) -> float:
        if not values:
            return 0.0
        if self.function == "sum":
            return sum(values)
        if self.function == "avg":
            return sum(values) / len(values)
        if self.function == "min":
            return min(values)
        if self.function == "max":
            return max(values)
        raise ReportError(f"unsupported aggregation: {self.function}")


@dataclass
class ReportConfig:
    title: str
    columns: list[ColumnSpec]
    aggregations: list[Aggregation] = field(default_factory=list)
    group_by: str | None = None
    filter_predicate: Any = None


def _read_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SourceNotFoundError(f"source file not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers: list[str] | None = None
    records: list[dict[str, Any]] = []
    for row in rows_iter:
        if headers is None:
            headers = [str(cell).strip() if cell is not None else "" for cell in row]
            continue
        if all(cell is None for cell in row):
            continue
        record = dict(zip(headers, row))
        records.append(record)
    workbook.close()
    logger.info("loaded %d rows from %s", len(records), path.name)
    return records


def _normalize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _apply_filter(records: list[dict[str, Any]], predicate: Any) -> list[dict[str, Any]]:
    if predicate is None:
        return records
    return [r for r in records if predicate(r)]


def _group_records(records: list[dict[str, Any]], config: ReportConfig) -> dict[Any, list[dict[str, Any]]]:
    if config.group_by is None:
        return {"__all__": records}
    groups: dict[Any, list[dict[str, Any]]] = {}
    for record in records:
        key = record.get(config.group_by)
        groups.setdefault(key, []).append(record)
    return groups


def _build_sheet(workbook: Workbook, config: ReportConfig, groups: dict[Any, list[dict[str, Any]]]) -> None:
    sheet = workbook.active
    sheet.title = config.title[:31]

    sheet.cell(row=1, column=1, value=config.title).font = Font(bold=True, size=14)
    header_row = 3
    for col_index, spec in enumerate(config.columns, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=spec.name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        sheet.column_dimensions[get_column_letter(col_index)].width = spec.width

    current_row = header_row + 1
    for group_key, records in sorted(groups.items(), key=lambda kv: str(kv[0])):
        for record in records:
            for col_index, spec in enumerate(config.columns, start=1):
                raw = record.get(spec.source_header)
                cell = sheet.cell(row=current_row, column=col_index, value=_normalize_value(raw))
                cell.border = THIN_BORDER
                if spec.number_format and isinstance(raw, (int, float)):
                    cell.number_format = spec.number_format
            current_row += 1
        if config.group_by is not None:
            current_row = _write_group_totals(sheet, config, records, group_key, current_row)

    if config.group_by is None and config.aggregations:
        current_row = _write_group_totals(sheet, config, records, "__all__", current_row)

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(config.columns))}{current_row - 1}"


def _write_group_totals(
    sheet: Any,
    config: ReportConfig,
    records: list[dict[str, Any]],
    group_key: Any,
    row: int,
) -> int:
    if not config.aggregations:
        return row
    label_cell = sheet.cell(row=row, column=1, value=f"{config.aggregations[0].label} — {group_key}")
    label_cell.font = Font(bold=True)
    label_cell.fill = TOTAL_FILL
    for agg in config.aggregations:
        col_index = next(
            (i for i, spec in enumerate(config.columns, start=1) if spec.source_header == agg.column),
            None,
        )
        if col_index is None:
            raise ColumnMismatchError(f"aggregation column missing: {agg.column}")
        values = [float(r.get(agg.column, 0) or 0) for r in records]
        cell = sheet.cell(row=row, column=col_index, value=agg.compute(values))
        spec = config.columns[col_index - 1]
        cell.number_format = spec.number_format or "#,##0.00"
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    return row + 1


def build_report(config: ReportConfig, sources: list[Path], output_path: Path) -> Path:
    records: list[dict[str, Any]] = []
    for source in sources:
        records.extend(_read_rows(source))
    records = _apply_filter(records, config.filter_predicate)
    if not records:
        raise ReportError("no data rows after filtering")

    groups = _group_records(records, config)
    workbook = Workbook()
    _build_sheet(workbook, config, groups)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    logger.info("report written to %s (%d rows)", output_path, len(records))
    return output_path
