from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from openpyxl import Workbook, load_workbook

from report_toolkit.core import Aggregation, ColumnSpec, ReportConfig, ReportError, build_report


def _make_source(path: Path, rows: list[tuple]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["region", "product", "revenue_usd", "units"])
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)
    return path


def test_build_report_with_group_totals(tmp_path: Path) -> None:
    source = _make_source(
        tmp_path / "sales.xlsx",
        [
            ("north", "widget", 120.5, 10),
            ("south", "gadget", 80.0, 4),
            ("north", "gadget", 200.0, 8),
        ],
    )
    config = ReportConfig(
        title="Q3 Sales",
        columns=[
            ColumnSpec(name="Region", source_header="region"),
            ColumnSpec(name="Product", source_header="product"),
            ColumnSpec(name="Revenue", source_header="revenue_usd", number_format="#,##0.00"),
        ],
        aggregations=[Aggregation(column="revenue_usd", function="sum")],
        group_by="region",
    )
    output = build_report(config, [source], tmp_path / "out.xlsx")
    assert output.exists()

    sheet = load_workbook(output).active
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "Q3 Sales" in values
    assert 320.5 in values
    assert 80.0 in values


def test_missing_source_raises(tmp_path: Path) -> None:
    import pytest

    config = ReportConfig(title="X", columns=[ColumnSpec("A", "a")])
    with pytest.raises(ReportError):
        build_report(config, [tmp_path / "nope.xlsx"], tmp_path / "o.xlsx")


def test_avg_aggregation(tmp_path: Path) -> None:
    source = _make_source(
        tmp_path / "s.xlsx",
        [("a", "x", 100, 1), ("b", "y", 300, 2)],
    )
    config = ReportConfig(
        title="Avg",
        columns=[ColumnSpec("Region", "region"), ColumnSpec("Amount", "revenue_usd")],
        aggregations=[Aggregation(column="revenue_usd", function="avg")],
    )
    output = build_report(config, [source], tmp_path / "avg.xlsx")
    values = [
        c.value
        for row in load_workbook(output).active.iter_rows()
        for c in row
        if isinstance(c.value, (int, float))
    ]
    assert any(abs(v - 200.0) < 0.01 for v in values)
